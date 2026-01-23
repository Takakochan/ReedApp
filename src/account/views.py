from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Count, Avg
from django.http import HttpResponse, JsonResponse
from reedsdata.models import Reedsdata
from .forms import ProfileUpdateForm
import csv
import json
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


@login_required
def account_view(request):
    """Main account dashboard"""
    user = request.user
    
    # Basic statistics
    total_reeds = Reedsdata.objects.filter(reedauthor=user).count()
    
    context = {
        'user': user,
        'total_reeds': total_reeds,
    }
    return render(request, 'account/account.html', context)


@login_required
def change_password_view(request):
    """Change user password"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Important to keep user logged in
            messages.success(request, 'Your password was successfully updated!')
            return redirect('account:account')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'account/change_password.html', {'form': form})


@login_required
def update_profile_view(request):
    """Update user profile information"""
    if request.method == 'POST':
        form = ProfileUpdateForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your profile was successfully updated!')
            return redirect('account:account')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = ProfileUpdateForm(instance=request.user)
    
    return render(request, 'account/update_profile.html', {'form': form})


@login_required
def account_statistics_view(request):
    """Display comprehensive account statistics with advanced analytics"""
    from .analytics import ReedAnalytics
    
    user = request.user
    reeds = Reedsdata.objects.filter(reedauthor=user)
    
    # Get selected instrument from request
    selected_instrument = request.GET.get('instrument')
    print(f"DEBUG: Selected instrument from request: '{selected_instrument}'")
    
    # Get selected parameters for correlation chart
    x_param = request.GET.get('x_param', 'hardness')
    y_param = request.GET.get('y_param', 'tone_color')
    print(f"DEBUG: Selected parameters - X: '{x_param}', Y: '{y_param}'")
    
    # Available parameters for dropdowns
    x_parameters = [
        ('hardness', 'Hardness'),
        ('chamber_temperature', 'Chamber Temperature'),
        ('chamber_humidity', 'Chamber Humidity'),
        ('harvest_year', 'Harvest Year'),
        ('gouging_machine', 'Gouging Machine'),
        ('profile_model', 'Profile Model'),
        ('diameter', 'Cane Diameter'),
        ('thickness', 'Thickness'),
        ('flexibility', 'Flexibility'),
        ('density', 'Density'),
        ('density_auto', 'Density Auto'),
        ('shaper', 'Shaper'),
        ('staple_model', 'Staple(ob)'),
        ('temperature', 'Temperature (from API)'),
        ('altitude', 'Altitude (from API)'),
        ('humidity', 'Humidity (from API)'),
        ('air_pressure', 'Air Pressure (from API)'),
        ('weather_description', 'Weather Description'),
    ]
    
    y_parameters = [
        ('tone_color', 'Tone Color'),
        ('intonation', 'Intonation'),
        ('playing_ease', 'Playing Ease'),
        ('response', 'Response'),
        ('latest_global_quality', 'Global Quality'),
    ]
    
    # Get available instruments for the user
    available_instruments = list(reeds.values_list('instrument', flat=True).distinct().order_by('instrument'))
    available_instruments = [instr for instr in available_instruments if instr]  # Remove None values
    print(f"DEBUG: Available instruments: {available_instruments}")
    
    # Basic counts
    total_reeds = reeds.count()
    
    # Basic statistics for backwards compatibility
    cane_brand_stats = reeds.values('cane_brand').annotate(
        count=Count('cane_brand')
    ).order_by('-count')[:5]
    
    quality_stats = {}
    if total_reeds > 0:
        quality_fields = ['stiffness', 'playing_ease', 'intonation', 'tone_color', 'response']
        for field in quality_fields:
            avg = reeds.aggregate(avg=Avg(field))['avg']
            if avg:
                quality_stats[field] = round(avg, 1)
    
    # Monthly reed creation stats (last 6 months)
    from django.utils import timezone
    from datetime import timedelta
    from django.db.models.functions import TruncMonth
    six_months_ago = timezone.now() - timedelta(days=180)
    monthly_stats = reeds.filter(date__gte=six_months_ago).annotate(
        month=TruncMonth('date')
    ).values('month').annotate(count=Count('id')).order_by('month')
    
    # Advanced analytics with instrument and parameter selection
    analytics = ReedAnalytics(user)
    advanced_stats = analytics.get_comprehensive_analysis(selected_instrument, x_param, y_param)
    
    context = {
        'total_reeds': total_reeds,
        'cane_brand_stats': cane_brand_stats,
        'quality_stats': quality_stats,
        'monthly_stats': monthly_stats,
        'advanced_analytics': advanced_stats,
        'has_sufficient_data': total_reeds >= 10,
        'available_instruments': available_instruments,
        'selected_instrument': selected_instrument,
        'x_parameters': x_parameters,
        'y_parameters': y_parameters,
        'selected_x_param': x_param,
        'selected_y_param': y_param,
    }
    return render(request, 'account/statistics.html', context)


@login_required
def delete_account_view(request):
    """Delete user account with confirmation"""
    if request.method == 'POST':
        password = request.POST.get('password')
        confirm_delete = request.POST.get('confirm_delete')

        if confirm_delete == 'DELETE' and request.user.check_password(password):
            # Delete user account (this will cascade delete all related data)
            request.user.delete()
            messages.success(request, 'Your account has been successfully deleted.')
            return redirect('home')
        else:
            if confirm_delete != 'DELETE':
                messages.error(request, 'You must type "DELETE" to confirm account deletion.')
            if not request.user.check_password(password):
                messages.error(request, 'Incorrect password.')

    return render(request, 'account/delete_account.html')


@login_required
def export_data_csv(request):
    """Export user's reed data as CSV"""
    # Get all reed data for the user
    reeds = Reedsdata.objects.filter(reedauthor=request.user).order_by('-date')

    # Create the HttpResponse with CSV content type
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="reedmanage_data_{datetime.now().strftime("%Y%m%d")}.csv"'

    # Create CSV writer
    writer = csv.writer(response)

    # Write header row
    header = [
        'Date', 'Instrument', 'Cane Brand', 'Harvest Year', 'Gouging Machine',
        'Profile Model', 'Diameter', 'Thickness', 'Hardness', 'Flexibility',
        'Density', 'Density Auto', 'Shaper', 'Staple Model', 'Thread Color',
        'Stiffness', 'Playing Ease', 'Intonation', 'Tone Color', 'Response',
        'Global Quality', 'Temperature', 'Humidity', 'Air Pressure', 'Altitude',
        'Weather Description', 'Location', 'Notes'
    ]
    writer.writerow(header)

    # Write data rows
    for reed in reeds:
        row = [
            reed.date.strftime('%Y-%m-%d') if reed.date else '',
            reed.instrument or '',
            reed.cane_brand or '',
            reed.harvest_year or '',
            reed.gouging_machine or '',
            reed.profile_model or '',
            reed.diameter or '',
            reed.thickness or '',
            reed.hardness or '',
            reed.flexibility or '',
            reed.density or '',
            reed.density_auto or '',
            reed.shaper or '',
            reed.staple_model or '',
            reed.thread_color or '',
            reed.stiffness or '',
            reed.playing_ease or '',
            reed.intonation or '',
            reed.tone_color or '',
            reed.response or '',
            reed.latest_global_quality or '',
            reed.temperature or '',
            reed.humidity or '',
            reed.air_pressure or '',
            reed.altitude or '',
            reed.weather_description or '',
            reed.location or '',
            reed.notes or '',
        ]
        writer.writerow(row)

    return response


@login_required
def export_data_excel(request):
    """Export user's reed data as Excel"""
    if not EXCEL_AVAILABLE:
        messages.error(request, 'Excel export is not available. Please try CSV export instead.')
        return redirect('account:account')

    # Get all reed data for the user
    reeds = Reedsdata.objects.filter(reedauthor=request.user).order_by('-date')

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Reed Data"

    # Style for header row
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Write header row
    headers = [
        'Date', 'Instrument', 'Cane Brand', 'Harvest Year', 'Gouging Machine',
        'Profile Model', 'Diameter', 'Thickness', 'Hardness', 'Flexibility',
        'Density', 'Density Auto', 'Shaper', 'Staple Model', 'Thread Color',
        'Stiffness', 'Playing Ease', 'Intonation', 'Tone Color', 'Response',
        'Global Quality', 'Temperature', 'Humidity', 'Air Pressure', 'Altitude',
        'Weather Description', 'Location', 'Notes'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Write data rows
    for row_num, reed in enumerate(reeds, 2):
        ws.cell(row=row_num, column=1, value=reed.date.strftime('%Y-%m-%d') if reed.date else '')
        ws.cell(row=row_num, column=2, value=reed.instrument or '')
        ws.cell(row=row_num, column=3, value=reed.cane_brand or '')
        ws.cell(row=row_num, column=4, value=reed.harvest_year or '')
        ws.cell(row=row_num, column=5, value=reed.gouging_machine or '')
        ws.cell(row=row_num, column=6, value=reed.profile_model or '')
        ws.cell(row=row_num, column=7, value=reed.diameter or '')
        ws.cell(row=row_num, column=8, value=reed.thickness or '')
        ws.cell(row=row_num, column=9, value=reed.hardness or '')
        ws.cell(row=row_num, column=10, value=reed.flexibility or '')
        ws.cell(row=row_num, column=11, value=reed.density or '')
        ws.cell(row=row_num, column=12, value=reed.density_auto or '')
        ws.cell(row=row_num, column=13, value=reed.shaper or '')
        ws.cell(row=row_num, column=14, value=reed.staple_model or '')
        ws.cell(row=row_num, column=15, value=reed.thread_color or '')
        ws.cell(row=row_num, column=16, value=reed.stiffness or '')
        ws.cell(row=row_num, column=17, value=reed.playing_ease or '')
        ws.cell(row=row_num, column=18, value=reed.intonation or '')
        ws.cell(row=row_num, column=19, value=reed.tone_color or '')
        ws.cell(row=row_num, column=20, value=reed.response or '')
        ws.cell(row=row_num, column=21, value=reed.latest_global_quality or '')
        ws.cell(row=row_num, column=22, value=reed.temperature or '')
        ws.cell(row=row_num, column=23, value=reed.humidity or '')
        ws.cell(row=row_num, column=24, value=reed.air_pressure or '')
        ws.cell(row=row_num, column=25, value=reed.altitude or '')
        ws.cell(row=row_num, column=26, value=reed.weather_description or '')
        ws.cell(row=row_num, column=27, value=reed.location or '')
        ws.cell(row=row_num, column=28, value=reed.notes or '')

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reedmanage_data_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)

    return response


@login_required
def export_data_json(request):
    """Export user's reed data as JSON"""
    # Get all reed data for the user
    reeds = Reedsdata.objects.filter(reedauthor=request.user).order_by('-date')

    # Build data structure
    data = {
        'export_date': datetime.now().isoformat(),
        'user': request.user.username,
        'total_reeds': reeds.count(),
        'reeds': []
    }

    for reed in reeds:
        reed_data = {
            'date': reed.date.strftime('%Y-%m-%d') if reed.date else None,
            'instrument': reed.instrument,
            'cane_brand': reed.cane_brand,
            'harvest_year': reed.harvest_year,
            'gouging_machine': reed.gouging_machine,
            'profile_model': reed.profile_model,
            'diameter': float(reed.diameter) if reed.diameter else None,
            'thickness': float(reed.thickness) if reed.thickness else None,
            'hardness': float(reed.hardness) if reed.hardness else None,
            'flexibility': float(reed.flexibility) if reed.flexibility else None,
            'density': float(reed.density) if reed.density else None,
            'density_auto': float(reed.density_auto) if reed.density_auto else None,
            'shaper': reed.shaper,
            'staple_model': reed.staple_model,
            'thread_color': reed.thread_color,
            'stiffness': float(reed.stiffness) if reed.stiffness else None,
            'playing_ease': float(reed.playing_ease) if reed.playing_ease else None,
            'intonation': float(reed.intonation) if reed.intonation else None,
            'tone_color': float(reed.tone_color) if reed.tone_color else None,
            'response': float(reed.response) if reed.response else None,
            'global_quality': float(reed.latest_global_quality) if reed.latest_global_quality else None,
            'temperature': float(reed.temperature) if reed.temperature else None,
            'humidity': float(reed.humidity) if reed.humidity else None,
            'air_pressure': float(reed.air_pressure) if reed.air_pressure else None,
            'altitude': float(reed.altitude) if reed.altitude else None,
            'weather_description': reed.weather_description,
            'location': reed.location,
            'notes': reed.notes,
        }
        data['reeds'].append(reed_data)

    # Create response
    response = HttpResponse(
        json.dumps(data, indent=2),
        content_type='application/json'
    )
    response['Content-Disposition'] = f'attachment; filename="reedmanage_data_{datetime.now().strftime("%Y%m%d")}.json"'

    return response